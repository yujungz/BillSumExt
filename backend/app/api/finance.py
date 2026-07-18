"""Finance API - supplier reconciliation, user statistics."""

import asyncio
import time
import uuid
from urllib.parse import quote

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import Response, FileResponse

from app.services import finance_service

router = APIRouter(prefix="/api/finance", tags=["finance"])


class FinanceQueryError(Exception):
    pass


# ── Common helpers ──

@router.get("/log-tables")
async def log_tables(site: str = Query(...)):
    return {"tables": await finance_service.get_log_tables(site)}


@router.get("/usernames")
async def usernames(
    site: str = Query(...),
    table: str = Query(""),
):
    return {"usernames": await finance_service.get_usernames(site, table)}


@router.get("/table-dates")
async def table_dates(table: str = Query(...)):
    start, end = finance_service.parse_log_table_dates(table)
    return {"start": start, "end": end}


# ── Supplier reconciliation ──

@router.get("/supplier")
async def supplier_query(
    site: str = Query(...),
    table: str = Query(...),
    username: str = Query(""),
    date_start: str = Query(""),
    date_end: str = Query(""),
    supplier_name: str = Query(""),
):
    try:
        rows = await finance_service.supplier_query(site, table, username, date_start, date_end, supplier_name)
        return {"rows": rows}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ── 供应商对账异步查询(全部用户数据量大) ──
_SUPPLIER_TASKS: dict[str, dict] = {}
_SUPPLIER_TASK_TTL = 3600


def _gc_supplier_tasks():
    now = time.time()
    expired = [k for k, v in _SUPPLIER_TASKS.items() if v["end"] and now - v["end"] > _SUPPLIER_TASK_TTL]
    for k in expired:
        _SUPPLIER_TASKS.pop(k, None)


@router.post("/supplier/query-async")
async def supplier_query_async(body: dict):
    _gc_supplier_tasks()
    task_id = uuid.uuid4().hex[:8]
    _SUPPLIER_TASKS[task_id] = {"status": "running", "result": None, "error": None,
                                "start": time.time(), "end": None}
    site = body.get("site", "")
    table = body.get("table", "")
    username = body.get("username", "")
    date_start = body.get("date_start", "")
    date_end = body.get("date_end", "")
    supplier_name = body.get("supplier_name", "")

    async def _run():
        try:
            rows = await finance_service.supplier_query(site, table, username, date_start, date_end, supplier_name)
            _SUPPLIER_TASKS[task_id]["result"] = rows
            _SUPPLIER_TASKS[task_id]["status"] = "done"
        except Exception as e:
            _SUPPLIER_TASKS[task_id]["status"] = "failed"
            _SUPPLIER_TASKS[task_id]["error"] = f"查询失败: {str(e)[:300]}"
        finally:
            _SUPPLIER_TASKS[task_id]["end"] = time.time()

    asyncio.create_task(_run())
    return {"task_id": task_id}


@router.get("/supplier/query-status")
async def supplier_query_status(task_id: str = Query(...)):
    t = _SUPPLIER_TASKS.get(task_id)
    if not t:
        raise HTTPException(404, detail="任务不存在或已过期")
    elapsed = (t["end"] or time.time()) - t["start"]
    return {"status": t["status"], "elapsed": round(elapsed, 1), "error": t["error"]}


@router.get("/supplier/query-result")
async def supplier_query_result(task_id: str = Query(...)):
    t = _SUPPLIER_TASKS.get(task_id)
    if not t:
        raise HTTPException(404, detail="任务不存在或已过期")
    if t["status"] != "done":
        raise HTTPException(400, detail="结果未就绪")
    rows = t["result"]
    _SUPPLIER_TASKS.pop(task_id, None)
    return {"rows": rows}


@router.get("/supplier/export")
async def supplier_export(
    site: str = Query(...),
    table: str = Query(...),
    username: str = Query(""),
    date_start: str = Query(""),
    date_end: str = Query(""),
    supplier_name: str = Query(""),
):
    rows = await finance_service.supplier_query(site, table, username, date_start, date_end, supplier_name)
    if not rows:
        raise HTTPException(400, detail="无数据可导出")
    # 构建 spec → 子进程生成 xlsx(不占 uvicorn GIL)
    from app.services.export_helper import generate_xlsx_subprocess
    col_keys = list(rows[0].keys())
    spec = {"sheets": [{"name": "供应商对账", "columns": [{"name": k, "label": k} for k in col_keys], "rows": rows}]}
    loop = asyncio.get_running_loop()
    xlsx_path = await generate_xlsx_subprocess(loop, spec)
    import os as _os
    ds = date_start.replace("-", "")
    de = date_end.replace("-", "")
    who = username or "全部"
    filename = f"supplier{ds}_{de}_{who}.xlsx"
    from fastapi.responses import FileResponse as _FR
    return _FR(path=xlsx_path, filename=filename,
               media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    ds = date_start.replace("-", "")
    de = date_end.replace("-", "")
    who = username or "全部"
    filename = f"supplier{ds}_{de}_{who}.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _build_supplier_excel(rows):
    import openpyxl
    from openpyxl.styles import Font, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "供应商对账"

    headers = list(rows[0].keys()) + ["折扣", "实付金额（USD）"]
    ws.append(headers)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")

    from app.services.excel_util import sanitize_row
    total_col = headers.index("总费用（USD）") + 1
    for r in rows:
        ws.append(sanitize_row(list(r.values()) + ["", ""]))

    ws.append(["" for _ in headers])
    total_row = ["" for _ in headers]
    total_row[0] = "合计"
    total_row[total_col - 1] = (
        f"=SUM({openpyxl.utils.get_column_letter(total_col)}2"
        f":{openpyxl.utils.get_column_letter(total_col)}{len(rows)+1})"
    )
    ws.append(total_row)

    from io import BytesIO

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── User statistics ──

@router.get("/user-stats")
async def user_stats(
    site: str = Query(...),
    table: str = Query(...),
    username: str = Query(...),
    date_start: str = Query(""),
    date_end: str = Query(""),
    granularity: str = Query(""),
    with_platform: bool = Query(False),
    with_total_cost: bool = Query(False),
    monthly_settle: bool = Query(False),
):
    try:
        glist = [x.strip() for x in granularity.split(",") if x.strip()] if granularity else []
        show_model = "model" in glist
        show_token = "token" in glist
        monthly = await finance_service.user_monthly(site, table, username, date_start, date_end,
                                                     show_model, with_platform, with_total_cost, monthly_settle)
        daily = await finance_service.user_daily(site, table, username, date_start, date_end, show_model, show_token)
        return {"monthly": monthly, "daily": daily}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


# ── 用户统计异步查询(大数据量) ──
_USER_STATS_QUERY_TASKS: dict[str, dict] = {}
_USER_STATS_QUERY_TTL = 3600


def _gc_user_stats_query_tasks():
    now = time.time()
    expired = [k for k, v in _USER_STATS_QUERY_TASKS.items() if v["end"] and now - v["end"] > _USER_STATS_QUERY_TTL]
    for k in expired:
        _USER_STATS_QUERY_TASKS.pop(k, None)


@router.post("/user-stats/query-async")
async def user_stats_query_async(body: dict):
    _gc_user_stats_query_tasks()
    task_id = uuid.uuid4().hex[:8]
    _USER_STATS_QUERY_TASKS[task_id] = {"status": "running", "result": None, "error": None,
                                        "start": time.time(), "end": None}
    site = body.get("site", "")
    table = body.get("table", "")
    username = body.get("username", "")
    date_start = body.get("date_start", "")
    date_end = body.get("date_end", "")
    granularity = body.get("granularity", "")
    with_platform = body.get("with_platform", False)
    with_total_cost = body.get("with_total_cost", False)
    monthly_settle = body.get("monthly_settle", False)

    async def _run():
        try:
            glist = [x.strip() for x in granularity.split(",") if x.strip()] if granularity else []
            show_model = "model" in glist
            show_token = "token" in glist
            monthly = await finance_service.user_monthly(site, table, username, date_start, date_end,
                                                         show_model, with_platform, with_total_cost, monthly_settle)
            daily = await finance_service.user_daily(site, table, username, date_start, date_end, show_model, show_token)
            _USER_STATS_QUERY_TASKS[task_id]["result"] = {"monthly": monthly, "daily": daily}
            _USER_STATS_QUERY_TASKS[task_id]["status"] = "done"
        except Exception as e:
            _USER_STATS_QUERY_TASKS[task_id]["status"] = "failed"
            _USER_STATS_QUERY_TASKS[task_id]["error"] = f"查询失败: {str(e)[:300]}"
        finally:
            _USER_STATS_QUERY_TASKS[task_id]["end"] = time.time()

    asyncio.create_task(_run())
    return {"task_id": task_id}


@router.get("/user-stats/query-status")
async def user_stats_query_status(task_id: str = Query(...)):
    t = _USER_STATS_QUERY_TASKS.get(task_id)
    if not t:
        raise HTTPException(404, detail="任务不存在或已过期")
    elapsed = (t["end"] or time.time()) - t["start"]
    return {"status": t["status"], "elapsed": round(elapsed, 1), "error": t["error"]}


@router.get("/user-stats/query-result")
async def user_stats_query_result(task_id: str = Query(...)):
    t = _USER_STATS_QUERY_TASKS.get(task_id)
    if not t:
        raise HTTPException(404, detail="任务不存在或已过期")
    if t["status"] != "done":
        raise HTTPException(400, detail="结果未就绪")
    result = t["result"]
    _USER_STATS_QUERY_TASKS.pop(task_id, None)
    return result


@router.get("/user-stats/detail")
async def user_stats_detail(
    site: str = Query(...),
    table: str = Query(...),
    username: str = Query(...),
    date_start: str = Query(""),
    date_end: str = Query(""),
    page: int = Query(1),
    size: int = Query(20),
):
    try:
        return await finance_service.user_detail(site, table, username, date_start, date_end, page, size)
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@router.get("/user-stats/export")
async def user_stats_export(
    site: str = Query(...),
    table: str = Query(...),
    username: str = Query(...),
    date_start: str = Query(""),
    date_end: str = Query(""),
    with_platform: str = Query("0"),
):
    monthly = await finance_service.user_monthly(site, table, username, date_start, date_end)
    daily = await finance_service.user_daily(site, table, username, date_start, date_end)
    detail_res = await finance_service.user_detail(site, table, username, date_start, date_end, page=1, page_size=1000000)
    detail = detail_res.get("data", [])
    if not monthly and not daily and not detail:
        raise HTTPException(400, detail="无数据可导出")
    content = _build_user_stats_excel(monthly, daily, detail, with_platform == "1")
    ds = date_start.replace("-", "")
    de = date_end.replace("-", "")
    filename = f"{username}_{ds}_{de}.xlsx"
    encoded = quote(filename)
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


@router.post("/user-stats/export-async")
async def user_stats_export_async(body: dict):
    site = body.get("site", "")
    table = body.get("table", "")
    username = body.get("username", "")
    date_start = body.get("date_start", "")
    date_end = body.get("date_end", "")
    with_platform = body.get("with_platform", False)
    with_detail = body.get("with_detail", True)
    with_total_cost = body.get("with_total_cost", True)
    monthly_settle = body.get("monthly_settle", False)
    granularity = body.get("granularity", "")
    glist = [x.strip() for x in granularity.split(",") if x.strip()] if granularity else []
    show_model = "model" in glist
    show_token = "token" in glist
    if not site or not table:
        raise HTTPException(400, detail="site, table 不能为空")
    task_id = finance_service.start_export_task(
        site, table, username, date_start, date_end, with_platform, with_detail,
        show_model, show_token, with_total_cost, monthly_settle
    )
    return {"task_id": task_id}


@router.get("/user-stats/export-status")
async def user_stats_export_status(task_id: str = Query(...)):
    status = finance_service.get_export_status(task_id)
    if not status:
        raise HTTPException(404, detail="任务不存在")
    return status


@router.get("/user-stats/export-download")
async def user_stats_export_download(task_id: str = Query(...)):
    import os
    status = finance_service.get_export_status(task_id)
    if not status:
        raise HTTPException(404, detail="任务不存在")
    if status["status"] != "done":
        raise HTTPException(400, detail="文件尚未生成完毕")
    file_path = status.get("file_path")
    if not file_path or not os.path.exists(file_path):
        raise HTTPException(500, detail="文件不存在")
    file_size = os.path.getsize(file_path)
    if file_size == 0:
        raise HTTPException(500, detail="生成的文件为空，请重试")
    username = status.get("username", "export")
    ds = status.get("date_start", "").replace("-", "")
    de = status.get("date_end", "").replace("-", "")
    filename = f"{username}_{ds}_{de}.xlsx" if ds else f"{username}.xlsx"
    encoded = quote(filename)
    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )
    """Remove specified columns from each row dict."""
    if not rows:
        return rows
    return [{k: v for k, v in r.items() if k not in strip_keys} for r in rows]


def _append_sheet_with_totals(ws, rows, total_fields):
    """Write data rows and append a SUM total row for specified fields."""
    if not rows:
        return
    import openpyxl
    from openpyxl.styles import Font, Alignment
    headers = list(rows[0].keys())
    ws.append(headers)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")
    from app.services.excel_util import sanitize_row
    for r in rows:
        ws.append(sanitize_row(list(r.values())))

    # Build total row with SUM formulas at correct column positions
    last_data_row = len(rows) + 1
    total_row = ["合计"] + ["" for _ in range(len(headers) - 1)]
    for field in total_fields:
        if field in headers:
            ci = headers.index(field) + 1
            cl = openpyxl.utils.get_column_letter(ci)
            total_row[ci - 1] = f"=SUM({cl}2:{cl}{last_data_row})"
    ws.append([])
    ws.append(total_row)


def _build_user_stats_excel(monthly, daily, detail=None, with_platform=False):
    import openpyxl
    from openpyxl.styles import Font, Alignment

    _PLATFORM_KEYS = {"平台额度"}

    if not with_platform:
        monthly = _strip_platform(monthly, _PLATFORM_KEYS)
        daily = _strip_platform(daily, _PLATFORM_KEYS)
        detail = _strip_platform(detail, _PLATFORM_KEYS)

    # Fields that should get SUM totals — numeric data columns, not IDs/names/unit-prices
    _NO_TOTAL_KEYS = {"结算周期", "用户名", "模型名", "用户ID", "Token名称", "日期", "时间", "序号",
                       "输入单价", "输出单价", "读取缓存单价", "创建缓存5M单价", "创建缓存1H单价"}

    def _total_fields(rows):
        if not rows:
            return []
        return [k for k in rows[0].keys() if k not in _NO_TOTAL_KEYS]

    wb = openpyxl.Workbook()
    ws_month = wb.active
    ws_month.title = "月汇总"
    if monthly:
        _append_sheet_with_totals(ws_month, monthly, _total_fields(monthly))

    ws_daily = wb.create_sheet("日统计")
    if daily:
        _append_sheet_with_totals(ws_daily, daily, _total_fields(daily))

    if detail:
        ws_detail = wb.create_sheet("用户明细")
        _append_sheet_with_totals(ws_detail, detail, _total_fields(detail))

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Site monthly report ──

@router.get("/site-report/preview")
async def site_report_preview(
    site: str = Query(...),
    table: str = Query(...),
    date_start: str = Query(""),
    date_end: str = Query(""),
    quota_type: str = Query("platform"),
):
    return await finance_service.site_report_preview(site, table, date_start, date_end, quota_type)


# ── 站点月报查询异步化(大数据量) ──
_SR_PREVIEW_TASKS: dict[str, dict] = {}
_SR_PREVIEW_TTL = 3600


def _gc_sr_preview_tasks():
    now = time.time()
    expired = [k for k, v in _SR_PREVIEW_TASKS.items() if v["end"] and now - v["end"] > _SR_PREVIEW_TTL]
    for k in expired:
        _SR_PREVIEW_TASKS.pop(k, None)


@router.post("/site-report/preview-async")
async def site_report_preview_async(body: dict):
    _gc_sr_preview_tasks()
    task_id = uuid.uuid4().hex[:8]
    _SR_PREVIEW_TASKS[task_id] = {"status": "running", "result": None, "error": None,
                                  "start": time.time(), "end": None}
    site = body.get("site", "")
    table = body.get("table", "")
    date_start = body.get("date_start", "")
    date_end = body.get("date_end", "")
    quota_type = body.get("quota_type", "platform")

    async def _run():
        try:
            result = await finance_service.site_report_preview(site, table, date_start, date_end, quota_type)
            _SR_PREVIEW_TASKS[task_id]["result"] = result
            _SR_PREVIEW_TASKS[task_id]["status"] = "done"
        except Exception as e:
            _SR_PREVIEW_TASKS[task_id]["status"] = "failed"
            _SR_PREVIEW_TASKS[task_id]["error"] = f"查询失败: {str(e)[:300]}"
        finally:
            _SR_PREVIEW_TASKS[task_id]["end"] = time.time()

    asyncio.create_task(_run())
    return {"task_id": task_id}


@router.get("/site-report/preview-status")
async def site_report_preview_status(task_id: str = Query(...)):
    t = _SR_PREVIEW_TASKS.get(task_id)
    if not t:
        raise HTTPException(404, detail="任务不存在或已过期")
    elapsed = (t["end"] or time.time()) - t["start"]
    return {"status": t["status"], "elapsed": round(elapsed, 1), "error": t["error"]}


@router.get("/site-report/preview-result")
async def site_report_preview_result(task_id: str = Query(...)):
    t = _SR_PREVIEW_TASKS.get(task_id)
    if not t:
        raise HTTPException(404, detail="任务不存在或已过期")
    if t["status"] != "done":
        raise HTTPException(400, detail="结果未就绪")
    result = t["result"]
    _SR_PREVIEW_TASKS.pop(task_id, None)
    return result


@router.post("/site-report/generate")
async def site_report_generate(body: dict):
    site = body.get("site")
    table = body.get("table")
    date_start = body.get("date_start")
    date_end = body.get("date_end")
    output_root = body.get("output_root", "E:/Workspaces/claude/BillSumExt/out")

    if not all([site, table, date_start, date_end]):
        raise HTTPException(400, detail="参数不完整")

    return await finance_service.generate_all_reports(
        site, table, date_start, date_end, output_root,
    )


@router.post("/site-report/generate-zip")
async def site_report_generate_zip(body: dict):
    site = body.get("site")
    table = body.get("table")
    date_start = body.get("date_start")
    date_end = body.get("date_end")

    if not all([site, table, date_start, date_end]):
        raise HTTPException(400, detail="参数不完整")

    zip_bytes = await finance_service.generate_reports_zip(site, table, date_start, date_end)
    ym = date_start.replace("-", "")[:6]
    filename = f"{site}_report{ym}.zip"
    return Response(
        content=zip_bytes,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── 站点月报异步生成(大数据量, 避免同步长连接超时) ──
_SR_ZIP_TASKS: dict[str, dict] = {}
_SR_ZIP_TTL = 3600


def _gc_sr_zip_tasks():
    now = time.time()
    expired = [k for k, v in _SR_ZIP_TASKS.items() if v["end"] and now - v["end"] > _SR_ZIP_TTL]
    for k in expired:
        _SR_ZIP_TASKS.pop(k, None)


@router.post("/site-report/generate-zip-async")
async def site_report_generate_zip_async(body: dict):
    _gc_sr_zip_tasks()
    site = body.get("site")
    table = body.get("table")
    date_start = body.get("date_start")
    date_end = body.get("date_end")
    if not all([site, table, date_start, date_end]):
        raise HTTPException(400, detail="参数不完整")
    task_id = uuid.uuid4().hex[:8]
    _SR_ZIP_TASKS[task_id] = {"status": "running", "zip": None, "error": None, "filename": None,
                              "start": time.time(), "end": None}

    async def _run():
        try:
            zip_bytes = await finance_service.generate_reports_zip(site, table, date_start, date_end)
            _SR_ZIP_TASKS[task_id]["zip"] = zip_bytes
            _SR_ZIP_TASKS[task_id]["filename"] = f"{site}_report{date_start.replace('-', '')[:6]}.zip"
            _SR_ZIP_TASKS[task_id]["status"] = "done"
        except Exception as e:
            _SR_ZIP_TASKS[task_id]["status"] = "failed"
            _SR_ZIP_TASKS[task_id]["error"] = f"生成失败: {str(e)[:300]}"
        finally:
            _SR_ZIP_TASKS[task_id]["end"] = time.time()

    asyncio.create_task(_run())
    return {"task_id": task_id}


@router.get("/site-report/generate-zip-status")
async def site_report_generate_zip_status(task_id: str = Query(...)):
    t = _SR_ZIP_TASKS.get(task_id)
    if not t:
        raise HTTPException(404, detail="任务不存在或已过期")
    elapsed = (t["end"] or time.time()) - t["start"]
    return {"status": t["status"], "elapsed": round(elapsed, 1), "error": t["error"]}


@router.get("/site-report/generate-zip-download")
async def site_report_generate_zip_download(task_id: str = Query(...)):
    t = _SR_ZIP_TASKS.get(task_id)
    if not t:
        raise HTTPException(404, detail="任务不存在或已过期")
    if t["status"] != "done":
        raise HTTPException(400, detail="文件尚未生成完毕")
    zip_bytes = t["zip"]
    filename = t["filename"] or "report.zip"
    _SR_ZIP_TASKS.pop(task_id, None)
    return Response(
        content=zip_bytes,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
