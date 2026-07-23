"""Statistics API."""

import asyncio
import io
import json
import logging
import os
import re
import subprocess
import sys
import tempfile
import time
import uuid
from pathlib import Path

from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import Response, FileResponse
from app.services.export_helper import generate_xlsx_subprocess

_XLSX_WORKER = os.path.join(os.path.dirname(os.path.dirname(__file__)), "services", "export_xlsx_worker.py")


def _col_letter(n):
    r = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        r = chr(65 + rem) + r
    return r


def _build_stats_row_xml(row_num, values):
    """拼一行 xlsx XML bytes(用于统计 sheet, 数据小)。"""
    import re as _re
    parts = [f'<row r="{row_num}">']
    for ci, val in enumerate(values):
        ref = f"{_col_letter(ci + 1)}{row_num}"
        if val is None or val == "":
            parts.append(f'<c r="{ref}"/>')
            continue
        if isinstance(val, (int, float)):
            v = str(int(val)) if isinstance(val, float) and val == int(val) else repr(val) if isinstance(val, float) else str(val)
            parts.append(f'<c r="{ref}"><v>{v}</v></c>')
        else:
            s = str(val).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")
            parts.append(f'<c r="{ref}" t="inlineStr"><is><t>{s}</t></is></c>')
    parts.append("</row>")
    return "".join(parts).encode("utf-8")


def _write_zip_metadata(zf, sheet_count, sheet_names):
    """写 xlsx 固定 XML(Content_Types, rels, workbook)。"""
    ct = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>',
    ]
    wbs = []
    wbrs = []
    for i in range(sheet_count):
        sn = sheet_names[i]
        ct.append(f'<Override PartName="/xl/worksheets/sheet{i+1}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>')
        wbs.append(f'<sheet name="{sn}" sheetId="{i+1}" r:id="rId{i+1}"/>')
        wbrs.append(f'<Relationship Id="rId{i+1}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{i+1}.xml"/>')
    ct.append("</Types>")
    zf.writestr("[Content_Types].xml", "".join(ct))
    zf.writestr("_rels/.rels",
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        '</Relationships>')
    zf.writestr("xl/workbook.xml",
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        f'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f'<sheets>{"".join(wbs)}</sheets></workbook>')
    zf.writestr("xl/_rels/workbook.xml.rels",
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f'{"".join(wbrs)}</Relationships>')


def _detail_where_sql(filters):
    """构建 WHERE 字符串(值内联, 供 mysql -e 子进程)。"""
    conditions = ["l.windup_type < 2"]
    if filters:
        for key, val in filters.items():
            if not val:
                continue
            safe_val = str(val).replace("\\", "\\\\").replace("'", "\\'")
            if key == 'date_start':
                conditions.append(f"l.created_at+28800 >= UNIX_TIMESTAMP('{safe_val} 00:00:00')")
            elif key == 'date_end':
                conditions.append(f"l.created_at+28800 <= UNIX_TIMESTAMP('{safe_val} 23:59:59')")
            elif re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', key):
                conditions.append(f"l.`{key}` LIKE '%{safe_val}%'")
    return " AND ".join(conditions)


def _dump_detail_tsv(mc, db_name, sql, tsv_path):
    """mysql 子进程: 执行 detail SQL → TSV 文件。"""
    cmd = [
        "mysql",
        f"--host={mc.host}", f"--port={mc.port}",
        f"--user={mc.user}", f"--password={mc.password}",
        "--skip-ssl", "--batch", "--quick",
        "-e", sql,
        db_name,
    ]
    with open(tsv_path, "wb") as fout:
        proc = subprocess.run(cmd, stdout=fout, stderr=subprocess.PIPE, timeout=3600)
    if proc.returncode != 0:
        raise RuntimeError(f"detail TSV dump failed: {proc.stderr.decode('utf-8', errors='replace')[:300]}")
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font

from app import database as db
from app.config import AppConfig
from app.services import stats_service
from app.services.excel_util import cell_value, sanitize_row

log = logging.getLogger(__name__)

log = logging.getLogger(__name__)

router = APIRouter(prefix="/api/stats", tags=["stats"])

# ── helpers shared by the original stats + the new detail sheet ──

FIELD_SQL = {
    "input_tokens": "l.prompt_tokens",
    "input_unit_price": "ROUND(l.model_ratio*2, 6)",
    "input_cost": "ROUND(l.group_ratio*l.model_ratio*2*l.prompt_tokens/1000000, 6)",
    "output_tokens": "l.completion_tokens",
    "output_unit_price": "ROUND(l.model_ratio*2*l.completion_ratio, 6)",
    "output_cost": "ROUND(l.group_ratio*l.model_ratio*2*l.completion_ratio*l.completion_tokens/1000000, 6)",
    "cache_read_tokens": "l.cache_tokens",
    "cache_read_unit_price": "ROUND(l.model_ratio*2*l.cache_ratio, 6)",
    "cache_read_cost": "ROUND(l.group_ratio*l.model_ratio*2*l.cache_ratio*l.cache_tokens/1000000, 6)",
    "cache_create_5m_tokens": "l.cache_creation_tokens_5m",
    "cache_create_5m_unit_price": "ROUND(l.model_ratio*2*1.25, 6)",
    "cache_create_5m_cost": "ROUND(l.group_ratio*l.model_ratio*2*1.25*l.cache_creation_tokens_5m/1000000, 6)",
    "cache_create_1h_tokens": "GREATEST(0, l.cache_creation_tokens - l.cache_creation_tokens_5m)",
    "cache_create_1h_unit_price": "ROUND(l.model_ratio*2*2.00, 6)",
    "cache_create_1h_cost": "ROUND(l.group_ratio*l.model_ratio*2*2.00*GREATEST(0, l.cache_creation_tokens - l.cache_creation_tokens_5m)/1000000, 6)",
    "cache_create_tokens": "GREATEST(l.cache_creation_tokens, l.cache_creation_tokens_5m)",
    "cache_create_cost": "ROUND(l.group_ratio*l.model_ratio*2*(1.25*l.cache_creation_tokens_5m + 2.00*GREATEST(0, l.cache_creation_tokens - l.cache_creation_tokens_5m))/1000000, 6)",
    "cache_total_tokens": "GREATEST(l.cache_creation_tokens, l.cache_creation_tokens_5m) + l.cache_tokens",
    "cache_total_cost": "ROUND(l.group_ratio*l.model_ratio*2*(1.25*l.cache_creation_tokens_5m + l.cache_ratio*l.cache_tokens + 2.00*GREATEST(0, l.cache_creation_tokens - l.cache_creation_tokens_5m))/1000000, 6)",
    "total_tokens": "GREATEST(l.cache_creation_tokens, l.cache_creation_tokens_5m) + l.cache_tokens + l.completion_tokens + l.prompt_tokens",
    "total_cost": "ROUND(l.group_ratio*l.model_ratio*2*(l.prompt_tokens + l.completion_ratio*l.completion_tokens + l.cache_ratio*l.cache_tokens + 1.25*l.cache_creation_tokens_5m + 2.00*GREATEST(0, l.cache_creation_tokens - l.cache_creation_tokens_5m))/1000000, 6)",
    "platform_quota": "l.quota*2/1000000",
}


def _col_width(label: str) -> float:
    w = len(label) * 1.8 + 2
    return max(w, 8)


def _detail_where(filters: dict | None):
    """Build WHERE clause for the detail SQL (no windup_type filter needed)."""
    conditions = ["l.windup_type < 2"]
    params = []
    if filters:
        ds = filters.get("date_start")
        de = filters.get("date_end")
        if ds:
            conditions.append("l.created_at+28800 >= UNIX_TIMESTAMP(%s)")
            params.append(f"{ds} 00:00:00")
        if de:
            conditions.append("l.created_at+28800 <= UNIX_TIMESTAMP(%s)")
            params.append(f"{de} 23:59:59")
    return " AND ".join(conditions), params


def _build_detail_columns(show_channel_name: bool, fields_json: str) -> list[dict]:
    """Build column definitions for the detail sheet.
    Returns list of {key, label} in display order."""
    cols = []
    # fixed: id, created_at, created_date
    cols.append({"key": "id", "label": "id"})
    cols.append({"key": "created_at", "label": "时间戳"})
    cols.append({"key": "created_date", "label": "创建日期"})
    # user fields
    cols.append({"key": "user_id", "label": "用户ID"})
    cols.append({"key": "username", "label": "用户"})
    cols.append({"key": "channel_id", "label": "渠道标识"})
    if show_channel_name:
        cols.append({"key": "channel_name", "label": "渠道名称"})
    cols.append({"key": "model_name", "label": "模型名称"})
    cols.append({"key": "token_id", "label": "Token标识"})
    cols.append({"key": "token_name", "label": "Token名称"})
    cols.append({"key": "group", "label": "分组"})
    cols.append({"key": "call_count", "label": "调用次数"})
    # dynamic fields from 字段选择
    if fields_json:
        try:
            flds = json.loads(fields_json)
            for x in flds:
                nm = x.get("name")
                if nm in FIELD_SQL:
                    cols.append({"key": nm, "label": x.get("label") or nm})
        except Exception:
            log.warning("failed to parse detail fields JSON", exc_info=True)
    return cols


def _build_detail_sql(table: str, detail_cols: list[dict], show_channel_name: bool,
                       where: str) -> str:
    """Build SELECT SQL for detail query (no ORDER BY/pagination; caller adds them)."""
    selects = []
    for c in detail_cols:
        k = c["key"]
        if k == "channel_name" and not show_channel_name:
            continue
        if k in FIELD_SQL:
            selects.append(f"{FIELD_SQL[k]} AS `{c['label']}`")
        elif k in _DETAIL_FIXED_SQL:
            selects.append(f"{_DETAIL_FIXED_SQL[k]} AS `{c['label']}`")
    return f"SELECT {', '.join(selects)} FROM `{table}` l WHERE {where}"


_DETAIL_FIXED_SQL = {
    "id": "l.id",
    "created_at": "l.created_at",
    "created_date": "FROM_UNIXTIME(l.created_at+28800)",
    "user_id": "l.user_id",
    "username": "l.username",
    "channel_id": "l.channel_id",
    "channel_name": "l.channel_name",
    "model_name": "l.model_name",
    "token_id": "l.token_id",
    "token_name": "l.token_name",
    "group": "l.`group`",
    "call_count": "1",
}

_CHUNK_SIZE = 50000


async def _write_detail_sheets(wb, config, detail_cols, where, params):
    """Streaming detail export with keyset pagination (no OFFSET).
    Uses `l.id < ?` which leverages the PRIMARY KEY index, O(n) regardless of depth."""
    from openpyxl.styles import Font
    app_config = AppConfig.load()
    db_name = config.db_name
    table = config.table_name

    # count (approximate; the loop stops naturally when no more rows)
    count_sql = f"SELECT COUNT(*) AS total FROM `{table}` l WHERE {where}"
    row = await db.fetch_one(count_sql, params, db=db_name)
    total = row["total"] if row else 0
    if not total:
        return

    headers = [c["label"] for c in detail_cols]

    # Build detail SQL once (no ORDER BY/pagination; keyset pagination added per batch)
    show_ch_name = any(c["key"] == "channel_name" for c in detail_cols)
    base_sql = _build_detail_sql(table, detail_cols, show_ch_name, where)
    LIMIT = 50000

    last_id = None
    sheet_idx = 0
    ws = None
    rows_in_sheet = 0
    processed = 0

    while processed < total:
        if last_id is None:
            sql = f"{base_sql} ORDER BY l.id DESC LIMIT {LIMIT}"
            chunk = await db.fetch_all(sql, params, db=db_name)
        else:
            sql = f"{base_sql} AND l.id < %s ORDER BY l.id DESC LIMIT {LIMIT}"
            chunk = await db.fetch_all(sql, params + [last_id], db=db_name)
        if not chunk:
            break

        for row_dict in chunk:
            if ws is None or rows_in_sheet >= 1000000:
                sheet_idx += 1
                name = "日志明细" if sheet_idx == 1 else f"日志明细_{sheet_idx}"
                ws = wb.create_sheet(name)
                rows_in_sheet = 0
                ws.append(sanitize_row(headers))

            vals = []
            for cdef in detail_cols:
                v = row_dict.get(cdef["label"])
                vals.append(float(v) if v is not None and hasattr(v, "__float__") else v)
            ws.append(sanitize_row(vals))
            rows_in_sheet += 1

        # Keyset: last row's id becomes the bound for the next batch
        last_id = chunk[-1].get("id")
        processed += len(chunk)

    log.info("detail export %s / %s total=%d written=%d", db_name, table, total, processed)


# ── Async detail export task machinery ──

DATA_DIR = Path(os.getenv("DATA_DIR", "/app/data"))
_STATS_EXPORT_TASKS: dict[str, dict] = {}


def start_stats_detail_task(site: str, table_name: str, filters: dict | None,
                            show_channel_name: bool, fields: str) -> str:
    """Start a background detail export task, return task_id immediately."""
    task_id = uuid.uuid4().hex[:8]
    _STATS_EXPORT_TASKS[task_id] = {
        "task_id": task_id,
        "status": "running",
        "progress": "排队中",
        "start_time": time.time(),
        "end_time": None,
        "file_path": None,
        "error": None,
        "site": site,
        "table_name": table_name,
    }

    async def _run():
        try:
            await _run_stats_detail(task_id, site, table_name, filters, show_channel_name, fields)
            _STATS_EXPORT_TASKS[task_id]["status"] = "done"
        except Exception as e:
            log.exception("stats detail export task %s failed", task_id)
            _STATS_EXPORT_TASKS[task_id]["status"] = "failed"
            _STATS_EXPORT_TASKS[task_id]["error"] = str(e)
        finally:
            _STATS_EXPORT_TASKS[task_id]["end_time"] = time.time()

    import asyncio
    asyncio.create_task(_run())
    return task_id


async def _run_stats_detail(task_id: str, site: str, table_name: str,
                            filters: dict | None,
                            show_channel_name: bool, fields: str):
    task = _STATS_EXPORT_TASKS[task_id]
    task["progress"] = "查询列配置..."
    show_ch = show_channel_name
    detail_cols = _build_detail_columns(show_ch, fields)
    if not detail_cols:
        task["error"] = "无可导出列"
        return

    where, params = _detail_where(filters)
    config_req = type("Config", (), {"db_name": f"sum_{site}", "table_name": table_name})()

    task["progress"] = "准备写入..."
    wb = openpyxl.Workbook(write_only=True)
    await _write_detail_sheets(wb, config_req, detail_cols, where, params)

    # Save to temp file
    out_dir = DATA_DIR / "stats_detail" / task_id
    out_dir.mkdir(parents=True, exist_ok=True)
    file_path = out_dir / "detail.xlsx"
    wb.save(str(file_path))

    task["file_path"] = str(file_path)
    task["progress"] = "完成"


def get_stats_detail_status(task_id: str) -> dict | None:
    t = _STATS_EXPORT_TASKS.get(task_id)
    if not t:
        return None
    elapsed = (t["end_time"] or time.time()) - t["start_time"]
    return {
        "task_id": task_id,
        "status": t["status"],
        "progress": t["progress"],
        "error": t.get("error"),
        "elapsed": round(elapsed, 1),
    }


def get_stats_detail_download(task_id: str) -> tuple[str, str] | None:
    t = _STATS_EXPORT_TASKS.get(task_id)
    if not t or t["status"] not in ("done", "failed"):
        return None
    path = t.get("file_path")
    if not path or not os.path.exists(path):
        return None
    fn = f"{t.get('site', 'stats')}_{t.get('table_name', 'detail')}_明细.xlsx"
    return path, fn


# ── Pydantic model ──


class StatsRequest(BaseModel):
    site: str
    table_name: str
    group_by: list[str] = []
    filters: dict | None = None
    show_zero: bool = True
    show_channel_name: bool = False
    show_log_detail: bool = False
    fields: str = ""


# ── Endpoints ──


@router.post("/query")
async def query_stats(req: StatsRequest):
    try:
        result = await stats_service.query_stats(
            req.site, req.table_name, req.group_by, req.filters, req.show_zero,
            show_channel_name=req.show_channel_name,
        )
        return {"data": result}
    except Exception as e:
        msg = str(e)
        # 常见大数据量错误给更友好的提示
        low = msg.lower()
        if "memory" in low or "tmp table" in low or "1046" in msg:
            raise HTTPException(500, detail=f"数据量过大，数据库临时表/内存不足。建议缩小统计范围或分时段查询。({msg[:150]})")
        if "lost connection" in low or "timeout" in low or "2013" in msg or "2006" in msg:
            raise HTTPException(500, detail=f"查询超时（数据量过大）。建议缩小统计范围或分时段查询。({msg[:150]})")
        raise HTTPException(500, detail=f"统计查询失败: {msg[:300]}")


# ── 异步统计查询(大数据量) ──
# 后台跑聚合，前端轮询进度，避免同步长连接超时导致"暂无数据"
_STATS_QUERY_TASKS: dict[str, dict] = {}
_STATS_QUERY_TTL = 3600  # 结果在内存保留 1 小时


@router.post("/query-async")
async def query_stats_async(req: StatsRequest):
    """启动后台统计查询，立即返回 task_id。"""
    _gc_stats_query_tasks()
    task_id = uuid.uuid4().hex[:8]
    _STATS_QUERY_TASKS[task_id] = {
        "status": "running", "result": None, "error": None,
        "start": time.time(), "end": None,
    }

    async def _run():
        try:
            result = await stats_service.query_stats(
                req.site, req.table_name, req.group_by, req.filters, req.show_zero,
                show_channel_name=req.show_channel_name,
            )
            _STATS_QUERY_TASKS[task_id]["result"] = result
            _STATS_QUERY_TASKS[task_id]["status"] = "done"
        except Exception as e:
            msg = str(e)
            _STATS_QUERY_TASKS[task_id]["status"] = "failed"
            _STATS_QUERY_TASKS[task_id]["error"] = f"统计查询失败: {msg[:300]}"
        finally:
            _STATS_QUERY_TASKS[task_id]["end"] = time.time()

    asyncio.create_task(_run())
    return {"task_id": task_id}


@router.get("/query-status")
async def query_stats_status(task_id: str = Query(...)):
    t = _STATS_QUERY_TASKS.get(task_id)
    if not t:
        raise HTTPException(404, detail="任务不存在或已过期")
    elapsed = (t["end"] or time.time()) - t["start"]
    return {"status": t["status"], "elapsed": round(elapsed, 1), "error": t["error"]}


@router.get("/query-result")
async def query_stats_result(task_id: str = Query(...)):
    t = _STATS_QUERY_TASKS.get(task_id)
    if not t:
        raise HTTPException(404, detail="任务不存在或已过期")
    if t["status"] != "done":
        raise HTTPException(400, detail="结果未就绪")
    rows = t["result"]
    # 取走即清，避免内存堆积
    _STATS_QUERY_TASKS.pop(task_id, None)
    return {"data": rows}


def _gc_stats_query_tasks():
    """清理超过 TTL 的任务，防内存泄漏。"""
    now = time.time()
    expired = [k for k, v in _STATS_QUERY_TASKS.items()
               if v["end"] and now - v["end"] > _STATS_QUERY_TTL]
    for k in expired:
        _STATS_QUERY_TASKS.pop(k, None)


@router.get("/distinct")
async def get_distinct(site: str = Query(...), table: str = Query(...), field: str = Query(...)):
    values = await stats_service.get_distinct_values(site, table, field)
    return {"values": values}


@router.post("/export")
async def export_stats(req: StatsRequest):
    result = await stats_service.query_stats(
        req.site, req.table_name, req.group_by, req.filters, req.show_zero,
        show_channel_name=req.show_channel_name,
    )
    if not result:
        return Response(content=b"", status_code=204)

    # ─── stats sheet ───
    col_def = [
        ("period_month", "月份"),
        ("period_day", "日期"),
        ("user_id", "用户ID"),
        ("username", "用户名"),
        ("channel_id", "渠道ID"),
        ("channel_name", "渠道"),
        ("model_name", "模型"),
        ("group_name", "分组"),
        ("token_name", "Token名称"),
        ("cn_buyer1", "采购员"),
        ("cn_supplier1", "供应商"),
        ("us_salesperson", "销售员"),
        ("call_count", "调用次数"),
        ("input_tokens", "输入token"),
        ("input_unit_price", "输入单价"),
        ("input_cost", "输入费用"),
        ("output_tokens", "输出token"),
        ("output_unit_price", "输出单价"),
        ("output_cost", "输出费用"),
        ("cache_read_tokens", "读取缓存token"),
        ("cache_read_unit_price", "读取缓存单价"),
        ("cache_read_cost", "读取缓存费用"),
        ("cache_create_5m_tokens", "创建缓存5M-token"),
        ("cache_create_5m_unit_price", "创建缓存5M单价"),
        ("cache_create_5m_cost", "创建缓存5M费用"),
        ("cache_create_1h_tokens", "创建缓存1H-token"),
        ("cache_create_1h_unit_price", "创建缓存1H单价"),
        ("cache_create_1h_cost", "创建缓存1H费用"),
        ("cache_create_tokens", "创建缓存token"),
        ("cache_create_cost", "创建缓存费用"),
        ("cache_total_tokens", "缓存总token"),
        ("cache_total_cost", "缓存总费用"),
        ("total_tokens", "总消耗token"),
        ("total_cost", "消费额度"),
        ("platform_quota", "平台额度"),
    ]
    visible = [(k, l) for k, l in col_def if any(r.get(k) is not None for r in result)]

    # 字段选择过滤：若前端传了 fields，只保留被选中的列（按前端传入顺序）
    if req.fields:
        try:
            flds = {x["name"] for x in json.loads(req.fields)}
            # col_def 的非可计算字段（如粒度列）始终保留，不在 fields 中但要保留
            # 粒度列属于固定列，不在 FIELD_SQL 中，保留
            data_keys = set(k for k, _ in col_def if k not in FIELD_SQL)
            visible = [(k, l) for k, l in visible if k in data_keys or k in flds]
        except Exception:
            pass

    if req.group_by:
        date_col = "period_day" if "day" in req.group_by else "period_month"
        result.sort(key=lambda r: (r.get(date_col) or ""), reverse=True)

    no_sum = {"period_month", "period_day", "user_id", "username",
              "channel_id", "channel_name", "model_name", "group_name",
              "token_name", "cn_buyer1", "cn_supplier1", "us_salesperson"}
    sum_cols = [(ik, k) for ik, (k, l) in enumerate(visible) if k not in no_sum]

    loop = asyncio.get_running_loop()
    content = await loop.run_in_executor(None, _build_stats_bytes, result, req.group_by, req.fields)
    filename = f"{req.site}_{req.table_name}_sum.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── 异步导出(大数据量, run_in_executor 不阻塞事件循环) ──
_STATS_EXPORT_TASKS: dict[str, dict] = {}
_STATS_EXPORT_TTL = 3600


def _gc_stats_export_tasks():
    now = time.time()
    expired = [k for k, v in _STATS_EXPORT_TASKS.items() if v["end"] and now - v["end"] > _STATS_EXPORT_TTL]
    for k in expired:
        _STATS_EXPORT_TASKS.pop(k, None)


def _build_stats_bytes(result, group_by, fields):
    """同步生成 stats xlsx bytes(write_only), 供 run_in_executor 调用。"""
    from openpyxl.cell import WriteOnlyCell
    col_def = [
        ("period_month", "月份"), ("period_day", "日期"), ("user_id", "用户ID"),
        ("username", "用户名"), ("channel_id", "渠道ID"), ("channel_name", "渠道"),
        ("model_name", "模型"), ("group_name", "分组"), ("token_name", "Token名称"),
        ("cn_buyer1", "采购员"), ("cn_supplier1", "供应商"), ("us_salesperson", "销售员"),
        ("call_count", "调用次数"), ("input_tokens", "输入token"), ("input_unit_price", "输入单价"),
        ("input_cost", "输入费用"), ("output_tokens", "输出token"), ("output_unit_price", "输出单价"),
        ("output_cost", "输出费用"), ("cache_read_tokens", "读取缓存token"),
        ("cache_read_unit_price", "读取缓存单价"), ("cache_read_cost", "读取缓存费用"),
        ("cache_create_5m_tokens", "创建缓存5M-token"), ("cache_create_5m_unit_price", "创建缓存5M单价"),
        ("cache_create_5m_cost", "创建缓存5M费用"), ("cache_create_1h_tokens", "创建缓存1H-token"),
        ("cache_create_1h_unit_price", "创建缓存1H单价"), ("cache_create_1h_cost", "创建缓存1H费用"),
        ("cache_create_tokens", "创建缓存token"), ("cache_create_cost", "创建缓存费用"),
        ("cache_total_tokens", "缓存总token"), ("cache_total_cost", "缓存总费用"),
        ("total_tokens", "总消耗token"), ("total_cost", "消费额度"), ("platform_quota", "平台额度"),
    ]
    visible = [(k, l) for k, l in col_def if any(r.get(k) is not None for r in result)]
    if fields:
        try:
            flds = {x["name"] for x in json.loads(fields)}
            data_keys = set(k for k, _ in col_def if k not in FIELD_SQL)
            visible = [(k, l) for k, l in visible if k in data_keys or k in flds]
        except Exception:
            pass
    if group_by:
        date_col = "period_day" if "day" in group_by else "period_month"
        result.sort(key=lambda r: (r.get(date_col) or ""), reverse=True)
    no_sum = {"period_month", "period_day", "user_id", "username", "channel_id",
              "channel_name", "model_name", "group_name", "token_name",
              "cn_buyer1", "cn_supplier1", "us_salesperson"}
    sum_cols = [(ik, k) for ik, (k, _) in enumerate(visible) if k not in no_sum]

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Stats")
    bold = Font(bold=True)
    header = []
    for _, label in visible:
        c = WriteOnlyCell(ws, value=label)
        c.font = bold
        header.append(c)
    ws.append(header)
    for row in result:
        vals = []
        for key, _ in visible:
            v = row.get(key)
            vals.append(cell_value(float(v) if v is not None and hasattr(v, "__float__") else v))
        ws.append(vals)
    total_row = ["合计"] + [""] * (len(visible) - 1)
    for ci, key in sum_cols:
        sm = sum(float(r.get(key)) for r in result if r.get(key) is not None and hasattr(r.get(key), "__float__"))
        if sm:
            total_row[ci] = sm
    ws.append([])
    ws.append(sanitize_row(total_row))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.post("/export-async")
async def export_stats_async(req: StatsRequest):
    _gc_stats_export_tasks()
    task_id = uuid.uuid4().hex[:8]
    _STATS_EXPORT_TASKS[task_id] = {"status": "running", "file_path": None, "error": None,
                                    "filename": "stats_export.xlsx",
                                    "start": time.time(), "end": None}

    async def _run():
        t = _STATS_EXPORT_TASKS[task_id]
        tmp_files = []
        try:
            result = await stats_service.query_stats(
                req.site, req.table_name, req.group_by, req.filters, req.show_zero,
                show_channel_name=req.show_channel_name,
            )
            if not result:
                t["status"] = "done"
                return

            # 文件名: 统计_站点_周期.xlsx (去掉 logs 前缀)
            period = req.table_name.replace("logs", "") if req.table_name.startswith("logs") else req.table_name
            fname = f"{req.site}_统计_{period}.xlsx"
            t["filename"] = fname

            # 构建 summary spec
            col_def = [
                ("period_month", "月份"), ("period_day", "日期"), ("user_id", "用户ID"),
                ("username", "用户名"), ("channel_id", "渠道ID"), ("channel_name", "渠道"),
                ("model_name", "模型"), ("group_name", "分组"), ("token_name", "Token名称"),
                ("cn_buyer1", "采购员"), ("cn_supplier1", "供应商"), ("us_salesperson", "销售员"),
                ("call_count", "调用次数"), ("input_tokens", "输入token"), ("input_unit_price", "输入单价"),
                ("input_cost", "输入费用"), ("output_tokens", "输出token"), ("output_unit_price", "输出单价"),
                ("output_cost", "输出费用"), ("cache_read_tokens", "读取缓存token"),
                ("cache_read_unit_price", "读取缓存单价"), ("cache_read_cost", "读取缓存费用"),
                ("cache_create_5m_tokens", "创建缓存5M-token"), ("cache_create_5m_unit_price", "创建缓存5M单价"),
                ("cache_create_5m_cost", "创建缓存5M费用"), ("cache_create_1h_tokens", "创建缓存1H-token"),
                ("cache_create_1h_unit_price", "创建缓存1H单价"), ("cache_create_1h_cost", "创建缓存1H费用"),
                ("cache_create_tokens", "创建缓存token"), ("cache_create_cost", "创建缓存费用"),
                ("cache_total_tokens", "缓存总token"), ("cache_total_cost", "缓存总费用"),
                ("total_tokens", "总消耗token"), ("total_cost", "消费额度"), ("platform_quota", "平台额度"),
            ]
            visible = [(k, l) for k, l in col_def if any(r.get(k) is not None for r in result)]
            if req.fields:
                try:
                    flds = {x["name"] for x in json.loads(req.fields)}
                    data_keys = set(k for k, _ in col_def if k not in FIELD_SQL)
                    visible = [(k, l) for k, l in visible if k in data_keys or k in flds]
                except Exception:
                    pass
            if req.group_by:
                dc = "period_day" if "day" in req.group_by else "period_month"
                result.sort(key=lambda r: (r.get(dc) or ""), reverse=True)

            summary_spec = {"sheets": [{
                "name": "数据统计",
                "columns": [{"name": k, "label": l} for k, l in visible],
                "rows": result,
            }]}

            loop = asyncio.get_running_loop()
            config = AppConfig.load()
            mc = config.mysql
            db_name = config.db_name(req.site)

            if req.show_log_detail:
                # ── 合并导出: 统计 sheet + 明细 sheets (全管道, 无 TSV 文件) ──
                _shm = "/dev/shm" if os.path.isdir("/dev/shm") else None
                xlsx_path = tempfile.mktemp(suffix=".xlsx", dir=_shm)

                # 构建 detail SQL(带中文别名, mysql --batch 输出即中文表头)
                detail_cols = _build_detail_columns(req.show_channel_name, req.fields)
                where = _detail_where_sql(req.filters)
                detail_sql = _build_detail_sql(req.table_name, detail_cols, req.show_channel_name, where)

                AWK_SCRIPT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                                          "services", "tsv_to_xlsx.awk")

                def _run_pipeline():
                    """mysql | awk | python zip — 零临时文件(仅 XML 在 /dev/shm)。"""
                    import zipfile as _zf
                    mysql_cmd = [
                        "mysql",
                        f"--host={mc.host}", f"--port={mc.port}",
                        f"--user={mc.user}", f"--password={mc.password}",
                        "--skip-ssl", "--batch", "--quick",
                        "-e", detail_sql,
                        db_name,
                    ]
                    awk_cmd = ["awk", "-f", AWK_SCRIPT,
                               "-v", "MAX_ROWS=1000000",
                               "-v", "SHEET_PREFIX=明细"]

                    mysql_proc = subprocess.Popen(mysql_cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                    awk_proc = subprocess.Popen(awk_cmd, stdin=mysql_proc.stdout,
                                                stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                    mysql_proc.stdout.close()

                    sheet_count = 0
                    sheet_names = []
                    detail_rows = 0
                    current_f = None
                    current_tmp = None

                    try:
                        with _zf.ZipFile(xlsx_path, "w", _zf.ZIP_DEFLATED) as zf:
                            # Phase 1: 统计 sheets(从内存 spec, 小数据, 先写入 ZIP)
                            for sspec in summary_spec.get("sheets", []):
                                sheet_count += 1
                                sname = sspec.get("name", f"Sheet{sheet_count}")
                                sheet_names.append(sname)
                                cols = sspec["columns"]
                                col_names = [c["name"] for c in cols]
                                col_labels = [c["label"] for c in cols]

                                tmp = tempfile.mktemp(suffix=".xml", dir=_shm)
                                with open(tmp, "wb") as wf:
                                    wf.write(b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>')
                                    wf.write(b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData>')
                                    wf.write(_build_stats_row_xml(1, col_labels))
                                    rn = 1
                                    for row in sspec.get("rows", []):
                                        rn += 1
                                        wf.write(_build_stats_row_xml(rn, [row.get(cn) for cn in col_names]))
                                    wf.write(b'</sheetData></worksheet>')
                                zf.write(tmp, f"xl/worksheets/sheet{sheet_count}.xml")
                                os.unlink(tmp)

                            # Phase 2: 明细 sheets(从 awk 管道, 流式写 ZIP)
                            for raw in awk_proc.stdout:
                                line = raw.rstrip(b"\n")

                                if line.startswith(b"SHEET_START:"):
                                    sheet_count += 1
                                    sname = line[len(b"SHEET_START:"):].decode("utf-8", "replace")
                                    sheet_names.append(sname)
                                    current_tmp = tempfile.mktemp(suffix=".xml", dir=_shm)
                                    current_f = open(current_tmp, "wb")
                                    current_f.write(b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>')
                                    current_f.write(b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData>')

                                elif line.startswith(b"SHEET_END"):
                                    if current_f:
                                        current_f.write(b'</sheetData></worksheet>')
                                        current_f.close()
                                        current_f = None
                                        zf.write(current_tmp, f"xl/worksheets/sheet{sheet_count}.xml")
                                        os.unlink(current_tmp)
                                        current_tmp = None

                                elif line.startswith(b"DONE:"):
                                    try:
                                        detail_rows = int(line[len(b"DONE:"):])
                                    except ValueError:
                                        pass

                                elif current_f:
                                    current_f.write(raw)

                            # Phase 3: 写固定 XML
                            _write_zip_metadata(zf, sheet_count, sheet_names)

                    finally:
                        if current_f:
                            current_f.close()
                        if current_tmp and os.path.exists(current_tmp):
                            os.unlink(current_tmp)
                        awk_proc.wait()
                        mysql_proc.wait()
                        if mysql_proc.returncode != 0:
                            err = mysql_proc.stderr.read().decode("utf-8", "replace")[:300]
                            raise RuntimeError(f"mysql failed: {err}")
                        if awk_proc.returncode != 0:
                            err = awk_proc.stderr.read().decode("utf-8", "replace")[:300]
                            raise RuntimeError(f"awk failed: {err}")

                    return detail_rows

                task["progress"] = "导出明细: 正在查询+生成..."
                detail_count = await loop.run_in_executor(None, _run_pipeline)

                if not os.path.exists(xlsx_path) or os.path.getsize(xlsx_path) == 0:
                    raise RuntimeError("xlsx 生成文件为空(0 字节)")
                t["file_path"] = xlsx_path
            else:
                # ── 仅统计 ──
                xlsx_path = await generate_xlsx_subprocess(loop, summary_spec)
                t["file_path"] = xlsx_path

            t["status"] = "done"
        except Exception as e:
            t["status"] = "failed"
            t["error"] = f"导出失败: {str(e)[:300]}"
        finally:
            # 清理临时文件(xlsx 除外)
            for tf in tmp_files:
                try:
                    os.unlink(tf)
                except OSError:
                    pass
            t["end"] = time.time()

    asyncio.create_task(_run())
    return {"task_id": task_id}


@router.get("/export-status")
async def export_stats_status(task_id: str = Query(...)):
    t = _STATS_EXPORT_TASKS.get(task_id)
    if not t:
        raise HTTPException(404, detail="任务不存在或已过期")
    elapsed = (t["end"] or time.time()) - t["start"]
    return {"status": t["status"], "elapsed": round(elapsed, 1), "error": t["error"]}


@router.get("/export-download")
async def export_stats_download(task_id: str = Query(...)):
    t = _STATS_EXPORT_TASKS.get(task_id)
    if not t:
        raise HTTPException(404, detail="任务不存在或已过期")
    if t["status"] != "done":
        raise HTTPException(400, detail="文件尚未生成完毕")
    file_path = t.get("file_path")
    filename = t.get("filename", "stats_export.xlsx")
    _STATS_EXPORT_TASKS.pop(task_id, None)
    if not file_path or not os.path.exists(file_path):
        raise HTTPException(500, detail="导出文件不存在")
    return FileResponse(path=file_path, filename=filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.post("/export-detail")
async def export_stats_detail(req: StatsRequest):
    """Export log detail as a separate xlsx file (chunked streaming)."""
    config_req = type("Config", (), {"db_name": f"sum_{req.site}", "table_name": req.table_name})()
    detail_cols = _build_detail_columns(req.show_channel_name, req.fields)
    if not detail_cols:
        return Response(content=b"", status_code=204)

    where, params = _detail_where(req.filters)

    wb = openpyxl.Workbook(write_only=True)
    await _write_detail_sheets(wb, config_req, detail_cols, where, params)

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"{req.site}_{req.table_name}_明细.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/export-detail-async")
async def export_stats_detail_async(req: StatsRequest):
    """Start background detail export and return task_id."""
    filters_dict = req.filters
    task_id = start_stats_detail_task(
        req.site, req.table_name, filters_dict,
        req.show_channel_name, req.fields
    )
    return {"task_id": task_id}


@router.get("/export-detail-status")
async def export_stats_detail_status(task_id: str = Query(...)):
    result = get_stats_detail_status(task_id)
    if not result:
        raise HTTPException(404, detail="任务不存在")
    return result


@router.get("/export-detail-download")
async def export_stats_detail_download(task_id: str = Query(...)):
    res = get_stats_detail_download(task_id)
    if not res:
        raise HTTPException(404, detail="明细文件尚未生成或已被清理")
    path, fn = res
    return FileResponse(path, filename=fn, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
