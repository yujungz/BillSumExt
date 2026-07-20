"""Query API - list, query, delete and export tables."""

import asyncio
import csv
import io
import json
import os
import re
import subprocess
import sys
import tempfile
import time
import uuid

from fastapi import APIRouter, Query, HTTPException, UploadFile, File
from fastapi.responses import Response, StreamingResponse, FileResponse

from app.config import AppConfig
from app.services import query_service, parser_service

router = APIRouter(prefix="/api/query", tags=["query"])

TABLE_NAME_RE = re.compile(r'^[a-zA-Z_][a-zA-Z0-9_]*$')


def _validate_table(name: str):
    if not TABLE_NAME_RE.match(name):
        raise HTTPException(400, detail=f"Invalid table name: {name}")


@router.get("/tables")
async def list_tables(site: str = Query(...), type: str = Query("raw")):
    if type == "raw":
        return {"tables": await query_service.list_raw_tables(site)}
    else:
        return {"tables": await query_service.list_output_tables(site)}


@router.get("/log-tables")
async def list_log_tables(site: str = Query(...)):
    return {"tables": await query_service.list_log_tables(site)}


@router.get("/columns")
async def get_columns(site: str = Query(...), table: str = Query(...)):
    _validate_table(table)
    return {"columns": await query_service.get_table_columns(site, table)}


@router.get("/data")
async def query_data(
    site: str = Query(...),
    table: str = Query(...),
    page: int = Query(1),
    size: int = Query(20),
    filters: str = Query(None),
    time_order: str = Query("desc"),
):
    _validate_table(table)
    import json
    f = json.loads(filters) if filters else None
    return await query_service.query_table(site, table, page, size, f, time_order)


@router.delete("/table")
async def delete_table(site: str = Query(...), table: str = Query(...)):
    _validate_table(table)
    await query_service.delete_table(site, table)
    return {"success": True, "message": f"Table {table} deleted"}


@router.get("/export")
async def export_table(
    site: str = Query(...),
    table: str = Query(...),
    format: str = Query("xlsx"),
    filters: str = Query(None),
    fields: str = Query(None),
):
    _validate_table(table)
    import json
    f = json.loads(filters) if filters else None
    fn = f"{site}_{table}"

    if format == "sql":
        content = query_service.export_table_sql(site, table)
        return Response(
            content=content,
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={fn}.sql"},
        )

    columns, rows = await query_service.export_all_data(site, table, f)
    cols = _apply_fields(columns, fields)

    if format == "csv":
        def _csv_stream():
            buf = io.StringIO()
            buf.write('﻿')
            writer = csv.writer(buf)
            writer.writerow([c['label'] for c in cols])
            for i, row in enumerate(rows):
                writer.writerow([str(row.get(c['name'], '')) for c in cols])
                if i % 10000 == 9999:  # 每 1 万行 flush 一次
                    yield buf.getvalue()
                    buf.seek(0)
                    buf.truncate(0)
            yield buf.getvalue()

        return StreamingResponse(
            _csv_stream(),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={fn}.csv"},
        )

    # xlsx - openpyxl write_only + 分sheet(每100万行一个sheet)
    content = _build_xlsx(cols, rows)
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fn}.xlsx"},
    )


# ── 异步导出(大数据量, 避免 uvicorn 同步 CPU 阻塞) ──

_EXPORT_TASKS: dict[str, dict] = {}
_EXPORT_TTL = 3600


def _gc_export_tasks():
    now = time.time()
    expired = [k for k, v in _EXPORT_TASKS.items() if v["end"] and now - v["end"] > _EXPORT_TTL]
    for k in expired:
        _EXPORT_TASKS.pop(k, None)


def _build_csv_bytes(cols, rows) -> bytes:
    """生成 CSV bytes(用于异步导出, 非流式)。"""
    buf = io.StringIO()
    buf.write('﻿')
    writer = csv.writer(buf)
    writer.writerow([c['label'] for c in cols])
    for row in rows:
        writer.writerow([str(row.get(c['name'], '')) for c in cols])
    return buf.getvalue().encode('utf-8')


def _export_csv_subprocess(mc, db_name, table, tmp_path, where=""):
    """TSV 导出: mysql 子进程直接写文件(完全绕过 Python GIL)。"""
    query = f"SELECT * FROM `{table}`{where}"
    cmd = [
        "mysql",
        f"--host={mc.host}", f"--port={mc.port}",
        f"--user={mc.user}", f"--password={mc.password}",
        "--skip-ssl", "--batch",
        "-e", query,
        db_name,
    ]
    with open(tmp_path, "wb") as fout:
        proc = subprocess.run(cmd, stdout=fout, stderr=subprocess.PIPE, timeout=3600)
    if proc.returncode != 0:
        raise RuntimeError(f"mysql export failed: {proc.stderr.decode('utf-8', errors='replace')[:300]}")


def _build_where_sql(filters):
    """从 filters dict 构建 WHERE 子句(用于 mysql -e 子进程, 参数化转义)。"""
    if not filters:
        return ""
    conditions = []
    for key, val in filters.items():
        if not val:
            continue
        safe_val = str(val).replace("\\", "\\\\").replace("'", "\\'")
        if key == 'created_at_start':
            conditions.append(f"created_at >= UNIX_TIMESTAMP('{safe_val} 00:00:00')-28800")
        elif key == 'created_at_end':
            conditions.append(f"created_at <= UNIX_TIMESTAMP('{safe_val} 23:59:59')-28800")
        elif re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', key):
            conditions.append(f"`{key}` LIKE '%{safe_val}%'")
    return (" WHERE " + " AND ".join(conditions)) if conditions else ""


# xlsx 转换子进程脚本路径
XLSX_WORKER = os.path.join(os.path.dirname(os.path.dirname(__file__)), "services", "export_xlsx_worker.py")


def _export_sql_subprocess(mc, db_name, table, tmp_path):
    """SQL 导出: mysqldump 子进程直接写文件。"""
    cmd = [
        "mysqldump",
        f"--host={mc.host}", f"--port={mc.port}",
        f"--user={mc.user}", f"--password={mc.password}",
        "--default-character-set=utf8mb4", "--skip-ssl",
        "--no-tablespaces", "--single-transaction", "--set-gtid-purged=OFF",
        db_name, table,
    ]
    with open(tmp_path, "wb") as fout:
        proc = subprocess.run(cmd, stdout=fout, stderr=subprocess.PIPE, timeout=3600)
    if proc.returncode != 0:
        raise RuntimeError(f"mysqldump failed: {proc.stderr.decode('utf-8', errors='replace')[:300]}")


def _export_xlsx_file(cols, rows, tmp_path):
    """xlsx 导出: openpyxl write_only 写文件(同步函数, 供 run_in_executor)。"""
    content = _build_xlsx(cols, rows)
    with open(tmp_path, "wb") as f:
        f.write(content)


def _run_xlsx_worker(tsv_path, xlsx_path, cols_json):
    """启动 xlsx 转换子进程(TSV → xlsx, 独立 Python 进程不占 uvicorn GIL)。"""
    proc = subprocess.run(
        [sys.executable, XLSX_WORKER, tsv_path, xlsx_path, cols_json],
        capture_output=True, timeout=3600,
    )
    return proc


@router.post("/export-async")
async def export_table_async(
    site: str = Query(...),
    table: str = Query(...),
    format: str = Query("xlsx"),
    filters: str = Query(None),
    fields: str = Query(None),
):
    _validate_table(table)
    _gc_export_tasks()
    import json
    f = json.loads(filters) if filters else None
    fn = f"{site}_{table}"

    task_id = uuid.uuid4().hex[:8]
    _EXPORT_TASKS[task_id] = {
        "status": "running", "file_path": None, "error": None,
        "media": "", "fn": fn, "format": format,
        "start": time.time(), "end": None,
    }

    async def _run():
        t = _EXPORT_TASKS[task_id]
        tmp_path = None
        try:
            config = AppConfig.load()
            db_name = config.db_name(site)
            mc = config.mysql
            loop = asyncio.get_running_loop()

            if format == "sql":
                # mysqldump 子进程 → 文件(完全不经过 Python)
                tmp_path = tempfile.NamedTemporaryFile(suffix=".sql", delete=False).name
                await loop.run_in_executor(None, _export_sql_subprocess, mc, db_name, table, tmp_path)
                t["media"] = "text/plain; charset=utf-8"
                t["fn"] = f"{fn}.sql"
            elif format == "csv":
                # mysql 子进程 → TSV 文件(完全不经过 Python)
                where = _build_where_sql(f)
                tmp_path = tempfile.NamedTemporaryFile(suffix=".tsv", delete=False).name
                await loop.run_in_executor(None, _export_csv_subprocess, mc, db_name, table, tmp_path, where)
                t["media"] = "text/csv; charset=utf-8"
                t["fn"] = f"{fn}.csv"
            else:
                # xlsx: 全管道 mysql | awk | zip(零临时文件, 最低磁盘 I/O)
                columns = await query_service.get_table_columns(site, table)
                cols = _apply_fields(columns, fields)

                # 构建 SELECT 带中文别名(mysql 直接输出列名作为 TSV header)
                col_select = ", ".join(f"`{c['name']}`" for c in cols)
                where = _build_where_sql(f)
                select_sql = f"SELECT {col_select} FROM `{table}`{where}"

                tmp_path = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name

                def _run_pipeline():
                    """mysql | awk | zip 管道, 零临时文件。"""
                    import zipfile as _zf
                    awk_script = os.path.join(
                        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                        "services", "tsv_to_xlsx.awk"
                    )
                    mysql_cmd = [
                        "mysql",
                        f"--host={mc.host}", f"--port={mc.port}",
                        f"--user={mc.user}", f"--password={mc.password}",
                        "--skip-ssl", "--batch",
                        "-e", select_sql,
                        db_name,
                    ]
                    awk_cmd = [
                        "awk", "-f", awk_script,
                        "-v", "MAX_ROWS=1000000",
                        "-v", "SHEET_PREFIX=Data",
                    ]

                    mysql_proc = subprocess.Popen(mysql_cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                    awk_proc = subprocess.Popen(awk_cmd, stdin=mysql_proc.stdout,
                                                stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                    mysql_proc.stdout.close()  # 让 mysql 能收 SIGPIPE

                    sheet_count = 0
                    sheet_names = []
                    global_rows = 0
                    tmp_dir = "/dev/shm" if os.path.isdir("/dev/shm") else None

                    try:
                        with _zf.ZipFile(tmp_path, "w", _zf.ZIP_DEFLATED) as zf:
                            current_f = None
                            current_tmp = None

                            for raw in awk_proc.stdout:
                                line = raw.rstrip(b"\n")

                                if line.startswith(b"SHEET_START:"):
                                    sheet_count += 1
                                    sname = line[len(b"SHEET_START:"):].decode("utf-8", "replace")
                                    sheet_names.append(sname)
                                    current_tmp = tempfile.mktemp(suffix=".xml", dir=tmp_dir)
                                    current_f = open(current_tmp, "wb")
                                    current_f.write(b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>')
                                    current_f.write(b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData>')

                                elif line.startswith(b"SHEET_END"):
                                    if current_f:
                                        current_f.write(b'</sheetData></worksheet>')
                                        current_f.close()
                                        current_f = None
                                        zf.write(current_tmp, f"xl/worksheets/sheet{sheet_count}.xml")
                                        os.unlink(current_tmp)
                                        current_tmp = None

                                elif line.startswith(b"DONE:"):
                                    try:
                                        global_rows = int(line[len(b"DONE:"):])
                                    except ValueError:
                                        pass

                                elif current_f:
                                    current_f.write(raw)

                            # 写固定 XML
                            ct = [
                                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                                '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
                                '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
                                '<Default Extension="xml" ContentType="application/xml"/>'
                                '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>',
                            ]
                            wbs = []
                            wbrs = []
                            for i in range(sheet_count):
                                sn = sheet_names[i]
                                ct.append(f'<Override PartName="/xl/worksheets/sheet{i+1}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>')
                                wbs.append(f'<sheet name="{sn}" sheetId="{i+1}" r:id="rId{i+1}"/>')
                                wbrs.append(f'<Relationship Id="rId{i+1}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{i+1}.xml"/>')
                            ct.append("</Types>")

                            zf.writestr("[Content_Types].xml", "".join(ct))
                            zf.writestr("_rels/.rels",
                                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                                '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
                                '</Relationships>')
                            zf.writestr("xl/workbook.xml",
                                f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                                f'<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                                f'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
                                f'<sheets>{"".join(wbs)}</sheets></workbook>')
                            zf.writestr("xl/_rels/workbook.xml.rels",
                                f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                                f'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                                f'{"".join(wbrs)}</Relationships>')

                    finally:
                        if current_f:
                            current_f.close()
                        if current_tmp and os.path.exists(current_tmp):
                            os.unlink(current_tmp)
                        awk_proc.wait()
                        mysql_proc.wait()
                        if mysql_proc.returncode != 0:
                            err = mysql_proc.stderr.read().decode("utf-8", "replace")[:300]
                            raise RuntimeError(f"mysql failed: {err}")
                        if awk_proc.returncode != 0:
                            err = awk_proc.stderr.read().decode("utf-8", "replace")[:300]
                            raise RuntimeError(f"awk failed: {err}")

                await loop.run_in_executor(None, _run_pipeline)

                if not os.path.exists(tmp_path) or os.path.getsize(tmp_path) == 0:
                    raise RuntimeError("xlsx 生成文件为空(0 字节)")
                t["media"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                t["fn"] = f"{fn}.xlsx"

            t["file_path"] = tmp_path
            t["status"] = "done"
        except Exception as e:
            t["status"] = "failed"
            t["error"] = f"导出失败: {str(e)[:300]}"
            if tmp_path:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass
        finally:
            t["end"] = time.time()

    asyncio.create_task(_run())
    return {"task_id": task_id}


@router.get("/export-status")
async def export_table_status(task_id: str = Query(...)):
    t = _EXPORT_TASKS.get(task_id)
    if not t:
        raise HTTPException(404, detail="任务不存在或已过期")
    elapsed = (t["end"] or time.time()) - t["start"]
    return {"status": t["status"], "elapsed": round(elapsed, 1), "error": t["error"]}


@router.get("/export-download")
async def export_table_download(task_id: str = Query(...)):
    t = _EXPORT_TASKS.get(task_id)
    if not t:
        raise HTTPException(404, detail="任务不存在或已过期")
    if t["status"] != "done":
        raise HTTPException(400, detail="文件尚未生成完毕")
    file_path = t.get("file_path")
    filename = t["fn"]
    media = t["media"]
    _EXPORT_TASKS.pop(task_id, None)
    if not file_path or not os.path.exists(file_path):
        raise HTTPException(500, detail="导出文件不存在")
    # FileResponse 流式发送文件, 不加载到内存
    return FileResponse(
        path=file_path,
        media_type=media,
        filename=filename,
    )


def _apply_fields(columns, fields_json):
    """Filter/rename columns per the field-selection config.
    fields_json: JSON '[{"name":..,"label":..}, ...]' (selected fields in order).
    None/invalid → all columns, label = field name.
    """
    default = [{"name": c["name"], "label": c["name"]} for c in columns]
    if not fields_json:
        return default
    import json
    try:
        flds = json.loads(fields_json)
    except Exception:
        return default
    db_names = {c["name"] for c in columns}
    out = []
    for x in flds:
        nm = x.get("name")
        if nm in db_names:
            out.append({"name": nm, "label": x.get("label") or nm})
    return out or default


def _build_xlsx(columns, rows):
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.cell import WriteOnlyCell
    from app.services.excel_util import cell_value

    wb = openpyxl.Workbook(write_only=True)
    bold = Font(bold=True)
    MAX_PER_SHEET = 1000000

    def _new_sheet(idx):
        ws = wb.create_sheet("Data" if idx == 1 else f"Data_{idx}")
        header = []
        for c in columns:
            cell = WriteOnlyCell(ws, value=c['label'])
            cell.font = bold
            header.append(cell)
        ws.append(header)
        return ws

    ws = _new_sheet(1)
    sheet_idx = 1
    row_in_sheet = 0
    for row in rows:
        if row_in_sheet >= MAX_PER_SHEET:
            sheet_idx += 1
            ws = _new_sheet(sheet_idx)
            row_in_sheet = 0
        ws.append([cell_value(row.get(c['name'], '')) for c in columns])
        row_in_sheet += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


COLUMN_ALIASES = {
    # channels / ex_channels
    "渠道ID": "id", "渠道名称": "name", "采购员": "buyer", "供应商": "supplier",
    "原始折扣": "discount_orig", "折扣": "discount", "原始折扣 折/刀": "discount_orig",
    # tokens / ex_tokens
    "TokenID": "id", "UserID": "user_id",
    # users / ex_users
    "用户ID": "id", "用户名称": "name", "用户名": "name",
    "备注": "remark", "销售人员": "seller", "销售员": "seller",
}


def _resolve_columns(raw_headers: list[str], db_columns: set[str]) -> list[str | None]:
    """Map file headers to db columns. Stops at first empty header to ignore trailing cols."""
    result: list[str | None] = []
    for h in raw_headers:
        h = h.strip()
        if not h:
            break  # stop at first empty column
        if h in db_columns:
            result.append(h)
            continue
        mapped = COLUMN_ALIASES.get(h)
        if mapped and mapped in db_columns:
            result.append(mapped)
        else:
            result.append(None)
    return result


def _map_row(row: tuple, col_map: list[str | None]) -> list | None:
    """Extract values from a row, only keeping columns that mapped to db fields."""
    vals = []
    for i, col in enumerate(col_map):
        if col is None:
            continue
        v = row[i] if i < len(row) else None
        vals.append(None if v is None else v)
    return vals if vals else None


def _map_row_csv(row: list[str], col_map: list[str | None]) -> list | None:
    """Extract values from a CSV row, only keeping columns that mapped to db fields."""
    vals = []
    for i, col in enumerate(col_map):
        if col is None:
            continue
        v = row[i] if i < len(row) else ""
        vals.append(None if v == "" else v)
    return vals if vals else None


async def _build_fillers(db, db_name: str, table: str, mapped_cols: list[str]) -> dict:
    """For NOT NULL columns missing from the import file, build a fill plan:
    {extra_col: (source_col, lookup_map)}. Currently fills ex_tokens.username
    via user_id -> users.id (取 users.username)，回退 ex_users.name；
    表不存在或无匹配则填空字符串。任何异常都不让导入失败。"""
    plan: dict = {}
    if table == 'ex_tokens' and 'username' not in mapped_cols and 'user_id' in mapped_cols:
        lookup: dict[str, str] = {}
        for sql, col in [("SELECT id, username FROM users", "username"),
                         ("SELECT id, name FROM ex_users", "name")]:
            try:
                rows = await db.fetch_all(sql, db=db_name)
                lookup = {str(r['id']): (r.get(col) or '') for r in rows}
                if lookup:
                    break
            except Exception:
                continue  # 表不存在(1146)等，尝试下一个来源
        plan['username'] = ('user_id', lookup)
    return plan


def _apply_fillers(vals: list, fillers: dict, src_idx: dict) -> None:
    """Append filler-derived values onto vals (in fillers dict order)."""
    for _ec, (src, lookup) in fillers.items():
        sv = vals[src_idx[src]] if src_idx[src] < len(vals) else None
        vals.append(lookup.get(str(sv), ''))


async def _not_null_defaults(db, db_name: str, table: str) -> dict:
    """Return {col: default} for NOT NULL columns: numeric→0, others→''.
    Used to fill empty cells so INSERT doesn't fail with 'X cannot be null'."""
    try:
        rows = await db.fetch_all(
            "SELECT COLUMN_NAME AS n, DATA_TYPE AS t FROM information_schema.COLUMNS "
            "WHERE TABLE_SCHEMA=%s AND TABLE_NAME=%s AND IS_NULLABLE='NO'",
            (db_name, table), db=db_name,
        )
    except Exception:
        return {}
    out: dict = {}
    for c in rows:
        dt = (c.get("t") or "").lower()
        out[c["n"]] = 0 if any(k in dt for k in ("int", "decimal", "double", "float")) else ""
    return out


def _fill_empty_with_defaults(vals: list, mapped_cols: list[str], defaults: dict) -> None:
    """In-place: replace None in vals with the column's NOT NULL default."""
    for i, col in enumerate(mapped_cols):
        if i < len(vals) and vals[i] is None and col in defaults:
            vals[i] = defaults[col]


@router.post("/import")
async def import_sql(site: str = Query(...), table: str = Query(...), overwrite: bool = Query(False), file: UploadFile = File(...)):
    _validate_table(table)
    config = AppConfig.load()
    db_name = config.db_name(site)

    filename = file.filename or "import.sql"
    ext = os.path.splitext(filename)[1].lower()

    content = await file.read()

    if ext == ".sql":
        mc = config.mysql
        suffix = ext
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False, mode="wb") as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        try:
            cmd = [
                "mysql",
                f"--host={mc.host}", f"--port={mc.port}",
                f"--user={mc.user}", f"--password={mc.password}",
                "--default-character-set=utf8mb4", "--skip-ssl",
                db_name,
            ]
            with open(tmp_path, "r", encoding="utf-8") as f:
                proc = subprocess.run(cmd, stdin=f, capture_output=True, text=True, timeout=600)
            if proc.returncode != 0:
                raise HTTPException(400, detail=f"Import failed: {proc.stderr[:500]}")
        finally:
            os.unlink(tmp_path)

    elif ext in (".xlsx", ".xls"):
        import openpyxl
        from app import database as db

        if overwrite:
            await db.execute(f"TRUNCATE TABLE `{table}`", db=db_name)

        # get actual db column names
        db_cols = await query_service.get_table_columns(site, table)
        db_col_set = {c["name"] for c in db_cols}

        buf = io.BytesIO(content)
        wb = openpyxl.load_workbook(buf, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        raw_headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
        col_map = _resolve_columns(raw_headers, db_col_set)  # list[str|None], same length as raw_headers
        mapped_cols = [c for c in col_map if c is not None]
        if not mapped_cols:
            raise HTTPException(400, detail="文件列名与表字段无法对应")

        # 补齐文件缺失的 NOT NULL 列（如 ex_tokens 缺 username 时按 user_id 关联 ex_users.name）
        fillers = await _build_fillers(db, db_name, table, mapped_cols)
        defaults = await _not_null_defaults(db, db_name, table)
        all_cols = mapped_cols + list(fillers.keys())
        src_idx = {src: mapped_cols.index(src) for src, _ in fillers.values()}

        cols_sql = ", ".join(f"`{c}`" for c in all_cols)
        placeholders = ", ".join(["%s"] * len(all_cols))
        # 追加模式：INSERT IGNORE，遇主键冲突的行跳过，不改变原有数据
        verb = "INSERT IGNORE INTO" if not overwrite else "INSERT INTO"
        sql = f"{verb} `{table}` ({cols_sql}) VALUES ({placeholders})"

        count = 0
        batch = []
        for row in rows_iter:
            vals = _map_row(row, col_map)
            if vals is not None:
                _fill_empty_with_defaults(vals, mapped_cols, defaults)
                _apply_fillers(vals, fillers, src_idx)
                batch.append(vals)
                if len(batch) >= 500:
                    await db.execute_many(sql, batch, db=db_name)
                    count += len(batch)
                    batch = []
        if batch:
            await db.execute_many(sql, batch, db=db_name)
            count += len(batch)
        wb.close()

    elif ext == ".csv":
        import csv as csv_mod
        from app import database as db

        if overwrite:
            await db.execute(f"TRUNCATE TABLE `{table}`", db=db_name)

        db_cols = await query_service.get_table_columns(site, table)
        db_col_set = {c["name"] for c in db_cols}

        text = content.decode("utf-8-sig")
        reader = csv_mod.reader(io.StringIO(text))
        raw_headers = [h.strip() for h in next(reader)]
        col_map = _resolve_columns(raw_headers, db_col_set)
        mapped_cols = [c for c in col_map if c is not None]
        if not mapped_cols:
            raise HTTPException(400, detail="文件列名与表字段无法对应")

        # 补齐文件缺失的 NOT NULL 列（如 ex_tokens 缺 username 时按 user_id 关联 ex_users.name）
        fillers = await _build_fillers(db, db_name, table, mapped_cols)
        defaults = await _not_null_defaults(db, db_name, table)
        all_cols = mapped_cols + list(fillers.keys())
        src_idx = {src: mapped_cols.index(src) for src, _ in fillers.values()}

        cols_sql = ", ".join(f"`{c}`" for c in all_cols)
        placeholders = ", ".join(["%s"] * len(all_cols))
        # 追加模式：INSERT IGNORE，遇主键冲突的行跳过，不改变原有数据
        verb = "INSERT IGNORE INTO" if not overwrite else "INSERT INTO"
        sql = f"{verb} `{table}` ({cols_sql}) VALUES ({placeholders})"

        count = 0
        batch = []
        for row in reader:
            vals = _map_row_csv(row, col_map)
            if vals is not None:
                _fill_empty_with_defaults(vals, mapped_cols, defaults)
                _apply_fillers(vals, fillers, src_idx)
                batch.append(vals)
                if len(batch) >= 500:
                    await db.execute_many(sql, batch, db=db_name)
                    count += len(batch)
                    batch = []
        if batch:
            await db.execute_many(sql, batch, db=db_name)
            count += len(batch)

    else:
        raise HTTPException(400, detail=f"不支持的文件格式: {ext}")

    return {"success": True, "message": "Import completed"}


@router.post("/parse")
async def parse_table(site: str = Query(...), table: str = Query(...)):
    if table not in parser_service.PARSEABLE_TABLES:
        raise HTTPException(400, detail=f"Table '{table}' is not parseable. Use: channels, tokens, users")

    parse_fn = {
        'channels': parser_service.parse_channels,
        'tokens': parser_service.parse_tokens,
        'users': parser_service.parse_users,
    }[table]

    try:
        result = await parse_fn(site)
        content = parser_service.build_excel_bytes(table, result["excel_headers"], result["excel_data"])
    except Exception as e:
        msg = str(e)
        # 1146 = 原始基础表(channels/users/tokens)未导入
        if "1146" in msg:
            raise HTTPException(400, detail=f"原始表 `{table}` 不存在。请先在『数据传输』中勾选基础表(channels/users/tokens)导入后，再执行拆解。")
        raise HTTPException(500, detail=f"拆解失败: {msg[:200]}")

    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={site}_ex_{table}.xlsx"},
    )
